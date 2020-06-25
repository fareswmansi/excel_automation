from openpyxl import workbook, load_workbook
from functions import checking_coordinates, get_cordinates, append_list, databse_loop, display_data, adding_letters, check_if_empty


#lists for data storage and automation processes
excel_file_1 = 'testme2.xlsx'
first_sheet = 'sheet1'
matched_strings = []
python_list = []
#!IMPORTANT! copy paste data from phpMyAdmin into this list to run program
database_list_of_lists = [('1', 33789190, 'Al Jasrah', '25', '1'),
                            ('2', 55860636, 'Al Jasrah', '50', '25.2841, 51.441'),
                            ('3', 55150250, 'Al Jasrah', '50', '25.2841, 51.441'),
                            ('4', 66570312, 'Al Jasrah', '24', '25.3318,51.5255'),
                            ('5', 55439821, 'Al Jasrah', '24', 'blah'),
                          ('5', 55512402, 'Al Jasrah', '24', 'blah')]
coordinates_list = []
just_testing = []

#start of user interface
print("Hello, welcome to the data transfer automation program. To proceed, enter YES")
first_choice = input("")

if (first_choice == 'YES' or first_choice == 'yes'):
    print("Please enter the name of the xlsx file you wish to automate.")
    name_of_sheet = input("")

    #authorization and opening of excel worksheets
    workbook = load_workbook(filename=name_of_sheet)
    sheet = workbook.active
    print("The following sheets are avaliable within the xlsx file: ")
    print(workbook.sheetnames)
    print("Which sheet would you like to access?")
    what_sheet = input("")

    #Asks what sheet you would like to access
    if (what_sheet == '1' or what_sheet == 'sheet1'):
        print("accessing batching_sheet....")

        print("To load data, enter 1. To skip that step and begin automating data, press 2.")
        load_or_automate = input("")

        #allows the user to simply see the data within the sheet without automating anything,
        #basically just adding functionality to the program
        if load_or_automate == '1':
            #variables defined at the start of script
            print("Data within the sheet: ")
            print(display_data())

        elif load_or_automate == '2':
            print("What is the day you wish to automate?")
            date_of_automation = input("")

            if date_of_automation >= '11':

                #loop through excel and database list to find similarities and append to matched_string list
                append_list(python_list)

                #loop through databse list and python list,
                #find similarities and put them in a seperate list (matched_strings)
                databse_loop(database_list_of_lists, python_list, matched_strings)

                #loop through excel spreadsheet and get coordinates of cell values
                #done in order to match cell value with cordinate
                #find cordinates, remove unwanted characters from returned string
                get_cordinates(matched_strings, coordinates_list)
                print(coordinates_list)

                #error catching, check if coordinates actually exist and match the strings
                checking_coordinates(coordinates_list)

                adding_letters(coordinates_list, just_testing)
                print(just_testing)

                check_if_empty(just_testing)

    elif what_sheet != '1':
        print("functionailty for other sheets has not been implemeneted yet. Thank you.")
        exit(0)

else:
    print("Thank you for using the data transfer automation program.")
    exit(0)
