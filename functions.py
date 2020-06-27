from openpyxl import workbook, load_workbook


excel_file_1 = 'testme2.xlsx'
workbook = load_workbook(filename='testme2.xlsx')
sheet = workbook.active

#display data to user interface
def display_data():
    for value in sheet.iter_rows(min_row=1,
                                 max_row=47,
                                 min_col=7,
                                 max_col=8,
                                 values_only=True):
        print(value)


#loop through excel and append values to python_list
def append_list(python_list):
    for value in sheet.iter_rows(min_row=25,
                                max_row=46,
                                min_col=6,
                                max_col=7,
                                values_only=True):
        python_list.append(value)

#loop through databse list and python list, append similarities to matched_string list
def databse_loop(database_list_of_lists, python_list, matched_strings):
    for list in database_list_of_lists:
        for phone_number in list:
            for second_list in python_list:
                for number in second_list:
                    if number == phone_number:
                        matched_strings.append(phone_number)

#get coordinates of matched_strings, append them to a seperate list and cut string to only
#have the coordinate rather than the whole returned tring
def get_cordinates(matched_strings, coordinates_list):
    for row in sheet.iter_rows(min_row=25,
                                           max_row=46,
                                           min_col=6,
                                           max_col=7):
        for cell in row:
            for number in matched_strings:
                if cell.value == number:
                    tryMe = cell
                    tryMe = str(tryMe).replace('<Cell \'sheet1\'.', '')
                    append_this = str(tryMe).replace('>', '')
                    coordinates_list.append(append_this)

#check coordinates existance within excel in order to catch errors
def checking_coordinates(coordinates_list):
    i = 0
    while i < len(coordinates_list):
        i += 1
        if i < len(coordinates_list):
            c = sheet[coordinates_list[i]]
            for row in sheet.iter_rows(min_row=25,
                                       max_row= 46,
                                       min_col=6,
                                       max_col=7):
                for cell in row:
                    if c == cell:
                        return True
        else:
            break

#change coordinate column to match input field
def adding_letters(coordinates_list, just_testing):
    i = 0
    firstTry = str(coordinates_list[0]).replace('G', 'M')
    just_testing.insert(0, firstTry)
    while i < len(coordinates_list):
        i += 1
        if i < len(coordinates_list):
            tryThis = str(coordinates_list[i]).replace('G', 'M')
            just_testing.append(tryThis)

#check if input field is empty
def check_if_empty(just_testing, add_to_these_coordinates):
    i = 0
    while i < len(just_testing):
        i += 1
        if i < len(just_testing):
            if sheet[just_testing[i]].value == None:
                add_to_these_coordinates.append(just_testing[i])

#match coordinates with list index
def match_coordinate_with_input(add_to_these_coordinates, database_list_of_lists, area_input_list, location_input_list, order_of_input_list):
    for coordinate in add_to_these_coordinates:
        indexme = str(coordinate).replace('M', 'G')
        i = 0
        while i < len(database_list_of_lists):
            i += 1
            if i < len(database_list_of_lists):
                if sheet[indexme].value == database_list_of_lists[i][1]:
                    order_of_input_list.append(database_list_of_lists[i][1])
                    area_input = str(database_list_of_lists[i][2]) + ' ' + str(database_list_of_lists[i][3])
                    location_input = database_list_of_lists[i][4]
                    area_input_list.append(area_input)
                    location_input_list.append(location_input)

#input area into matched string index in excel
"""def area_input_excel(area_input_list, order_of_input_list, add_to_these_coordinates, final_coordinate_list):
    sheet[add_to_these_coordinates[0]] = area_input_list[0]
    def testing_loop(order_of_input_list, add_to_these_coordinates, area_input_list):
        i = 0
        while i < len(order_of_input_list):
            i += 1
            if i < len(order_of_input_list):
                sheet[add_to_these_coordinates[i]] = area_input_list[i]
                for coordinate in add_to_these_coordinates:
                    final_index = coordinate.replace('M', 'Q')
                    final_coordinate_list.append(final_index)"""


def area_input_excel(area_input_list, add_to_these_coordinates, finial_coordinate_list, order_of_input_list):
    sheet[add_to_these_coordinates[0]] = area_input_list[0]
    def testing_loop(order_of_input_list, add_to_these_coordinates, area_input_list):
        i = 0
        while i < len(order_of_input_list):
            i += 1
            if i < len(order_of_input_list):
                sheet[add_to_these_coordinates[i]] = area_input_list[i]
                for coordinate in add_to_these_coordinates:
                    final_index = coordinate.replace('M', 'Q')
                    finial_coordinate_list.append(final_index)

#input location into matched string index in excel
def location_input_excel(location_input_list, final_coordinate_list, order_of_input_list):
    sheet[final_coordinate_list[0]] = location_input_list[0]
    i = 0
    while i < len(order_of_input_list):
        i += 1
        if i < len(order_of_input_list):
            sheet[final_coordinate_list[i]] = location_input_list[i]
            workbook.save('testme2.xlsx')

"""def area_input_excel(area_input_list, add_to_these_coordinates):
    for coordinate in add_to_these_coordinates:
        for area in area_input_list:
            sheet[coordinate] = area

def location_input_excel(location_input_list, final_coordinate_list, add_to_these_coordinates):
    for coordinate in add_to_these_coordinates:
        final_index = coordinate.replace('M', 'Q')
        final_coordinate_list.append(final_index)
        for index in final_coordinate_list:
            for location in location_input_list:
                sheet[index] = location
                workbook.save('testme2.xlsx')"""
